#!/usr/bin/env python3
"""
Script de procesamiento del Excel de diagnóstico de Ibeas de Juarros.
Genera los archivos JSON en src/data/ a partir de:
  matriz_ibeas_datos_exhaustiva_v4_todos_los_datos_OE.xlsx

Uso:
  python3 scripts/procesar_excel.py

Requiere: openpyxl
  pip install openpyxl
"""

import json
import sys
import os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
EXCEL_PATH = BASE_DIR / "matriz_ibeas_datos_exhaustiva_v4_todos_los_datos_OE.xlsx"
DATA_DIR = BASE_DIR / "src" / "data"

# ─────────────────────────────────────────────────────────────
def load_sheet(wb, name):
    """Carga una hoja como lista de dicts usando la primera fila como cabecera."""
    ws = wb[name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        rows.append(row)
    return rows

def clean(val):
    """Limpia valores None y tipos para JSON."""
    if val is None:
        return None
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        # Redondear a 4 decimales como máximo
        if val == int(val):
            return int(val)
        return round(val, 4)
    return val

def parse_oes_relacionados(txt):
    """Convierte 'OE 2, OE 6' → [2, 6]"""
    if not txt:
        return []
    result = []
    for part in str(txt).split(","):
        part = part.strip()
        if part.startswith("OE "):
            try:
                result.append(int(part[3:].strip()))
            except ValueError:
                pass
    return result

def parse_oe_principal(txt):
    """Convierte 'OE 4' → 4"""
    if not txt:
        return None
    txt = str(txt).strip()
    if txt.startswith("OE "):
        try:
            return int(txt[3:].strip())
        except ValueError:
            pass
    return None

# ─────────────────────────────────────────────────────────────
def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: No se encuentra el archivo Excel en {EXCEL_PATH}")
        sys.exit(1)

    print(f"Abriendo: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    print(f"Hojas: {wb.sheetnames}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # ── 1. Indicadores completos (de Matriz_OE_AUE) ──────────────
    print("\n[1/5] Procesando Matriz_OE_AUE…")
    raw = load_sheet(wb, "Matriz_OE_AUE")
    indicadores = []
    for row in raw:
        if not row.get("indicador"):
            continue
        oe_flags = {i: bool(row.get(f"OE_{i}")) for i in range(1, 11)}
        oe_rel = parse_oes_relacionados(row.get("OEs_relacionados"))
        ind = {
            "id": clean(row.get("id_registro")),
            "orden": clean(row.get("orden_fuente")),
            "tipo_registro": row.get("tipo_registro"),
            "ambito": row.get("ambito"),
            "grupo_tematico": row.get("grupo_tematico"),
            "bloque": row.get("bloque"),
            "indicador": row.get("indicador"),
            "valor_ibeas": clean(row.get("valor_ibeas")),
            "valor_espana": clean(row.get("valor_espana")),
            "unidad": row.get("unidad"),
            "mediana_referencia": clean(row.get("mediana_referencia")),
            "q1": clean(row.get("q1")),
            "q3": clean(row.get("q3")),
            "z_iqr": clean(row.get("z_iqr")),
            "etiqueta_num": clean(row.get("etiqueta_num")),
            "tipo_hallazgo": row.get("tipo_hallazgo_peor_v3"),
            "signo": row.get("signo_peor_v3"),
            "z_score": clean(row.get("z_peor_v3")),
            "disponibilidad": row.get("disponibilidad_valor_ibeas"),
            "disponible": str(row.get("disponibilidad_valor_ibeas", "")).strip().lower() == "disponible",
            "observaciones": row.get("observaciones"),
            "fuente": row.get("fuente"),
            "archivo_origen": row.get("archivo_origen"),
            "hoja_origen": row.get("hoja_origen"),
            "oe_principal": parse_oe_principal(row.get("OE_principal")),
            "titulo_oe_principal": row.get("titulo_OE_principal"),
            "oes_relacionados": oe_rel,
            "oe_flags": oe_flags,
        }
        indicadores.append(ind)

    print(f"  → {len(indicadores)} indicadores procesados")

    # ── 2. Vista por OE (de Todos_los_datos_OE) ──────────────────
    print("[2/5] Procesando Todos_los_datos_OE…")
    raw_oe = load_sheet(wb, "Todos_los_datos_OE")
    por_oe = {i: [] for i in range(1, 11)}
    for row in raw_oe:
        oe_txt = str(row.get("OE_despliegue", "")).strip()
        oe_num = parse_oe_principal(oe_txt)
        if oe_num and 1 <= oe_num <= 10:
            rec = {
                "id": clean(row.get("id_registro")),
                "tipo_vinculo": row.get("tipo_vinculo_OE"),
                "grupo_tematico": row.get("grupo_tematico"),
                "bloque": row.get("bloque"),
                "indicador": row.get("indicador"),
                "valor_ibeas": clean(row.get("valor_ibeas")),
                "valor_espana": clean(row.get("valor_espana")),
                "unidad": row.get("unidad"),
                "mediana_referencia": clean(row.get("mediana_referencia")),
                "etiqueta_num": clean(row.get("etiqueta_num")),
                "tipo_hallazgo": row.get("tipo_hallazgo_peor_v3"),
                "disponibilidad": row.get("disponibilidad_valor_ibeas"),
                "disponible": str(row.get("disponibilidad_valor_ibeas", "")).strip().lower() == "disponible",
                "fuente": row.get("fuente"),
                "observaciones": row.get("observaciones"),
                "oe_principal": parse_oe_principal(row.get("OE_principal")),
                "oes_relacionados": parse_oes_relacionados(row.get("OEs_relacionados")),
            }
            if rec.get("indicador"):
                por_oe[oe_num].append(rec)

    for oe_num, items in por_oe.items():
        print(f"  OE {oe_num}: {len(items)} registros")

    # ── 3. Vacíos / cobertura ─────────────────────────────────────
    print("[3/5] Procesando Cobertura_y_vacios…")
    raw_vacios = load_sheet(wb, "Cobertura_y_vacios")
    vacios = []
    for row in raw_vacios:
        if row.get("indicador"):
            vacios.append({
                "grupo_tematico": row.get("grupo_tematico"),
                "bloque": row.get("bloque"),
                "indicador": row.get("indicador"),
                "archivo_origen": row.get("archivo_origen"),
                "motivo": row.get("motivo_o_cautela"),
            })
    print(f"  → {len(vacios)} vacíos identificados")

    # ── 4. Resumen por OE ─────────────────────────────────────────
    print("[4/5] Procesando Resumen_OE…")
    raw_resumen = load_sheet(wb, "Resumen_OE")
    resumen_oe = {}
    for row in raw_resumen:
        oe_txt = str(row.get("OE", "")).strip()
        oe_num = parse_oe_principal(oe_txt)
        if oe_num:
            resumen_oe[str(oe_num)] = {
                "oe": oe_num,
                "titulo": row.get("Título"),
                "total": clean(row.get("Registros asignados")),
                "disponibles": clean(row.get("Con dato municipal disponible")),
                "no_disponibles": clean(row.get("Sin dato / N.D.")),
                "grupos_tematicos": row.get("Grupos temáticos presentes"),
            }

    # ── 5. Estadísticas globales ───────────────────────────────────
    print("[5/5] Calculando estadísticas globales…")
    total = len(indicadores)
    disponibles = sum(1 for i in indicadores if i["disponible"])
    no_disponibles = total - disponibles
    grupos = sorted(list(set(i["grupo_tematico"] for i in indicadores if i["grupo_tematico"])))
    fuentes_unicas = sorted(list(set(i["fuente"] for i in indicadores if i["fuente"])))
    tipos_hallazgo = sorted(list(set(i["tipo_hallazgo"] for i in indicadores if i["tipo_hallazgo"])))

    resumen_global = {
        "municipio": "Ibeas de Juarros",
        "total_indicadores": total,
        "disponibles": disponibles,
        "no_disponibles": no_disponibles,
        "tasa_disponibilidad_pct": round(100 * disponibles / total, 1) if total else 0,
        "por_oe": {str(k): len(v) for k, v in por_oe.items()},
        "resumen_oe": resumen_oe,
        "grupos_tematicos": grupos,
        "fuentes": fuentes_unicas,
        "tipos_hallazgo": tipos_hallazgo,
        "total_vacios": len(vacios),
        "generado": datetime.now().isoformat(),
        "archivo_excel": EXCEL_PATH.name,
    }

    # ── Escritura de archivos ──────────────────────────────────────
    def write_json(path, data, label):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        size = os.path.getsize(path)
        print(f"  ✓ {path.name} ({size // 1024} KB)")

    print("\nEscribiendo JSON en src/data/:")
    write_json(DATA_DIR / "indicadores.json", indicadores, "indicadores")
    write_json(DATA_DIR / "por_oe.json", por_oe, "por_oe")
    write_json(DATA_DIR / "vacios.json", vacios, "vacios")
    write_json(DATA_DIR / "resumen.json", resumen_global, "resumen")

    # Actualizar municipio.json con datos reales
    municipio = {
        "_placeholder": False,
        "nombre": "Ibeas de Juarros",
        "codigo_ine": "09178",
        "provincia": "Burgos",
        "comunidad_autonoma": "Castilla y León",
        "comarca": "Comarca del Arlanzón",
        "resumen_estadistico": resumen_global,
    }
    write_json(DATA_DIR / "municipio.json", municipio, "municipio")

    print(f"\n✅ Procesamiento completado · {total} indicadores · {len(vacios)} vacíos")
    print(f"   Disponibilidad: {disponibles}/{total} ({resumen_global['tasa_disponibilidad_pct']}%)")
    for oe_num, items in por_oe.items():
        r = resumen_oe.get(str(oe_num), {})
        print(f"   OE {oe_num}: {len(items)} registros ({r.get('disponibles','?')} disponibles)")

if __name__ == "__main__":
    main()
